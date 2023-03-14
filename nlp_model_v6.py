import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.feature_selection import chi2
import numpy as np
from sklearn.naive_bayes import MultinomialNB
from openpyxl import load_workbook

def read_excel_in_batches(filename, batch_size):
    wb = load_workbook(filename=filename)
    sheet_names = wb.sheetnames
    
    dfs = []
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        max_row = sheet.max_row
        
        for i in range(1, max_row, batch_size):
            df = pd.DataFrame(sheet.iter_rows(min_row=i, max_row=min(i+batch_size-1, max_row), values_only=True), columns=[cell.value for cell in sheet[1]])
            df = df.dropna(subset=['Consumer_complaint_narrative'])
            dfs.append(df[['Product', 'Consumer_complaint_narrative']])
    
    return pd.concat(dfs)
df_chunks = read_excel_in_batches(r'C:\Users\M_Thiruveedula\Downloads\complaints_v1.xlsx', 10000)
#df = pd.concat(df_chunks)
#df = pd.concat([chunk for chunk in df_chunks])
#df = pd.concat([chunk[col] for chunk in df_chunks if pd.notnull(chunk['Consumer_complaint_narrative'])])

col = ['Product', 'Consumer_complaint_narrative']
df = pd.concat([chunk[col] for chunk in df_chunks if pd.notnull(chunk['Consumer_complaint_narrative'])])

#df = pd.concat([chunk[col] for chunk in df_chunks if pd.notnull(chunk['Consumer complaint narrative'])])
#df = pd.concat([chunk[col] for chunk in df_chunks if pd.notnull(chunk['Consumer complaint narrative'])], axis=0)


df.columns = ['Product', 'Consumer_complaint_narrative']

#df = df.drop(df[df['Product'] == 'Credit card or prepaid card'].index)

category_id_df = df[['Product']].drop_duplicates().reset_index(drop=True)
category_id_df['category_id'] = category_id_df.index
category_to_id = dict(category_id_df.values)
id_to_category = dict(category_id_df[['category_id', 'Product']].values)

tfidf = TfidfVectorizer(sublinear_tf=True, min_df=5, norm='l2', encoding='latin-1', ngram_range=(1, 2), stop_words='english')

def batch_generator(df, batch_size):
    while True:
        for i in range(0, len(df), batch_size):
            batch = df.iloc[i:i+batch_size]
            features = tfidf.transform(batch['Consumer_complaint_narrative']).toarray()
            labels = batch['Product'].map(category_to_id).values
            yield features, labels

BATCH_SIZE = 1000
train_batches = batch_generator(df, BATCH_SIZE)
test_batches = batch_generator(df, BATCH_SIZE)

clf = MultinomialNB()

for i in range(10):  # Train for 10 epochs
    print(f"Epoch {i+1}")
    for j, (X_batch, y_batch) in enumerate(train_batches):
        clf.partial_fit(X_batch, y_batch, classes=np.array(list(category_to_id.values())))
        if j % 10 == 0:
            print(f"Processed {j*BATCH_SIZE} samples")
    
    # Evaluate on test set after each epoch
    test_accuracy = 0
    num_test_samples = 0
    for X_batch, y_batch in test_batches:
        test_accuracy += np.sum(clf.predict(X_batch) == y_batch)
        num_test_samples += len(y_batch)
    test_accuracy /= num_test_samples
    print(f"Test accuracy: {test_accuracy:.4f}")
